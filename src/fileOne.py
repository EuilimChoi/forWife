from openpyxl import Workbook
from openpyxl import load_workbook

def fileOne():
    workbook = load_workbook("./excelfilses/testfile1.xlsx")
    worksheet = workbook["mysheet1"]
    cellvalue = worksheet["A1"].value

    targetworkbook = load_workbook("./excelfilses/testfile10.xlsx")
    targetsheet = targetworkbook["mysheet10"]
    targetsheet["A1"] = cellvalue

    targetworkbook.save("./excelfilses/testfile10.xlsx")

    print('fileOne copy')
    return 'fileOne copy'





