from openpyxl import Workbook

def createFiles(files):
    for num in range(1, files):
        wb = Workbook()
        mysheet = wb.create_sheet(f"mysheet{num}")
        mysheet.cell(num, 1).value = f"fileInfo{num}"
        wb.save(f"./excelfilses/testfile{num}.xlsx")
    return f'{files} files created'





